"""
CJK text utilities for OOXML processing.

Provides:
- Display width calculation (full-width vs half-width characters)
- Korean language attribute injection for OOXML XML files
  - DrawingML (a:rPr) for PPTX/XLSX
  - WordprocessingML (w:rPr → w:lang) for DOCX
- CJK font injection for both DrawingML and WordML namespaces
- WCAG 2.1 color contrast checking
- CJK text-box width estimation
- EMU ↔ pixel conversion helpers
"""

from __future__ import annotations

import glob
import os
import unicodedata
from typing import Optional
from xml.etree import ElementTree as ET

import defusedxml.minidom


def get_display_width(text: str) -> int:
    """Return the display width of *text*, counting full-width chars as 2.

    Uses the Unicode East Asian Width property:
    - F (Full-width) and W (Wide): 2 units  (CJK ideographs, hangul, etc.)
    - Everything else: 1 unit

    >>> get_display_width("Hello")
    5
    >>> get_display_width("한글")
    4
    >>> get_display_width("Hello 세계")
    10
    """
    if not text:
        return 0
    width = 0
    for ch in str(text):
        eaw = unicodedata.east_asian_width(ch)
        width += 2 if eaw in ("F", "W") else 1
    return width


def auto_fit_columns(
    ws,
    padding: int = 2,
    min_width: int = 8,
    max_width: int = 60,
) -> None:
    """Adjust column widths of an openpyxl worksheet using CJK-aware width.

    Args:
        ws: openpyxl Worksheet object.
        padding: Extra character units to add.
        min_width: Minimum column width.
        max_width: Maximum column width.
    """
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):
        best = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            lines = str(val).split("\n")
            best = max(best, max(get_display_width(line) for line in lines))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(best + padding, max_width)


# ---------------------------------------------------------------------------
# DrawingML Korean lang injection (a:rPr — for PPTX/XLSX)
# ---------------------------------------------------------------------------


def inject_korean_lang(xml_dir: str, lang: str = "ko-KR", alt_lang: str = "en-US") -> int:
    """Add lang/altLang attributes to ``<a:rPr>`` elements missing them.

    Uses DOM parsing (defusedxml.minidom) instead of regex to safely skip
    XML comments, CDATA sections, and other non-element content.

    Operates on all ``*.xml`` files under *xml_dir* (recursive).
    Returns the number of files modified.

    Intended for post-processing unpacked PPTX or XLSX packages where
    PptxGenJS / docx-js did not set language attributes.
    """
    count = 0
    for path in glob.glob(os.path.join(xml_dir, "**", "*.xml"), recursive=True):
        with open(path, "r", encoding="utf-8") as fh:
            original = fh.read()
        try:
            dom = defusedxml.minidom.parseString(original.encode("utf-8"))
        except Exception:
            continue  # Skip malformed XML

        modified = False
        for elem in dom.getElementsByTagName("a:rPr"):
            if not elem.getAttribute("lang"):
                elem.setAttribute("lang", lang)
                elem.setAttribute("altLang", alt_lang)
                modified = True

        if modified:
            with open(path, "wb") as fh:
                fh.write(dom.toxml(encoding="UTF-8"))
            count += 1
    return count


# ---------------------------------------------------------------------------
# WordprocessingML Korean lang injection (w:rPr → w:lang — for DOCX)
# ---------------------------------------------------------------------------

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def inject_korean_lang_word(
    xml_dir: str,
    lang: str = "ko-KR",
    east_asia: str = "ko-KR",
    bidi: str = "ar-SA",
) -> int:
    """Add w:lang elements to w:rPr in Word XML files.

    Unlike ``inject_korean_lang()`` which targets DrawingML ``a:rPr`` (PPTX),
    this function targets WordprocessingML ``w:r`` runs and ensures each has
    a ``<w:lang>`` child inside ``<w:rPr>``.

    If a ``w:r`` has no ``w:rPr``, one is created as the first child
    (OOXML spec requires rPr to be the first child of w:r).

    Operates on all ``*.xml`` files under *xml_dir* (recursive).
    Returns the number of files modified.
    """
    count = 0
    for path in glob.glob(os.path.join(xml_dir, "**", "*.xml"), recursive=True):
        with open(path, "r", encoding="utf-8") as fh:
            original = fh.read()
        try:
            dom = defusedxml.minidom.parseString(original.encode("utf-8"))
        except Exception:
            continue

        modified = False

        # Process every w:r (run) element
        for run in dom.getElementsByTagNameNS(W_NS, "r"):
            # Find or create w:rPr
            rpr = None
            for child in run.childNodes:
                if (
                    getattr(child, "localName", None) == "rPr"
                    and child.namespaceURI == W_NS
                ):
                    rpr = child
                    break

            if rpr is None:
                # Create w:rPr as first child of w:r
                rpr = dom.createElementNS(W_NS, "w:rPr")
                if run.firstChild:
                    run.insertBefore(rpr, run.firstChild)
                else:
                    run.appendChild(rpr)

            # Check if w:lang already exists inside this rPr
            has_lang = any(
                getattr(c, "localName", None) == "lang" and c.namespaceURI == W_NS
                for c in rpr.childNodes
            )
            if has_lang:
                continue

            lang_elem = dom.createElementNS(W_NS, "w:lang")
            lang_elem.setAttribute("w:val", lang)
            lang_elem.setAttribute("w:eastAsia", east_asia)
            lang_elem.setAttribute("w:bidi", bidi)
            rpr.appendChild(lang_elem)
            modified = True

        if modified:
            with open(path, "wb") as fh:
                fh.write(dom.toxml(encoding="UTF-8"))
            count += 1
    return count


# ---------------------------------------------------------------------------
# CJK font injection (dual namespace: WordML + DrawingML)
# ---------------------------------------------------------------------------

_DML_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

_DEFAULT_CJK_FONTS = {
    "eastAsia": "Malgun Gothic",
    "hAnsi": "Malgun Gothic",
}


def inject_cjk_fonts(
    xml_dir: str,
    fonts: dict[str, str] | None = None,
    *,
    target: str = "auto",
) -> int:
    """Inject CJK font declarations into OOXML XML files.

    Supports two namespaces selected by *target*:
    - ``"word"``   → ``w:rFonts`` inside ``w:rPr`` (DOCX body text)
    - ``"drawing"`` → ``a:ea`` inside ``a:rPr`` (PPTX / XLSX charts & shapes)
    - ``"auto"``   → detect from directory structure (``word/`` → word, ``ppt/`` or ``xl/`` → drawing)

    *fonts* is a dict of attribute-name → font-name mappings. Defaults to
    ``{"eastAsia": "Malgun Gothic", "hAnsi": "Malgun Gothic"}``.

    Returns the number of files modified.
    """
    fonts = fonts or dict(_DEFAULT_CJK_FONTS)

    if target == "auto":
        target = _detect_font_target(xml_dir)

    if target == "word":
        return _inject_cjk_fonts_word(xml_dir, fonts)
    else:
        return _inject_cjk_fonts_drawing(xml_dir, fonts)


def _detect_font_target(xml_dir: str) -> str:
    """Detect whether xml_dir is a Word or Drawing package."""
    if os.path.isdir(os.path.join(xml_dir, "word")):
        return "word"
    return "drawing"


def _inject_cjk_fonts_word(xml_dir: str, fonts: dict[str, str]) -> int:
    """Add w:rFonts attributes to w:rPr elements in Word XML."""
    count = 0
    for path in glob.glob(os.path.join(xml_dir, "**", "*.xml"), recursive=True):
        with open(path, "r", encoding="utf-8") as fh:
            original = fh.read()
        try:
            dom = defusedxml.minidom.parseString(original.encode("utf-8"))
        except Exception:
            continue

        modified = False
        for rpr in dom.getElementsByTagNameNS(W_NS, "rPr"):
            # Find or create w:rFonts
            rfonts = None
            for child in rpr.childNodes:
                if (
                    getattr(child, "localName", None) == "rFonts"
                    and child.namespaceURI == W_NS
                ):
                    rfonts = child
                    break

            if rfonts is None:
                rfonts = dom.createElementNS(W_NS, "w:rFonts")
                if rpr.firstChild:
                    rpr.insertBefore(rfonts, rpr.firstChild)
                else:
                    rpr.appendChild(rfonts)

            for attr, font in fonts.items():
                w_attr = f"w:{attr}"
                if not rfonts.getAttribute(w_attr):
                    rfonts.setAttribute(w_attr, font)
                    modified = True

        if modified:
            with open(path, "wb") as fh:
                fh.write(dom.toxml(encoding="UTF-8"))
            count += 1
    return count


def _inject_cjk_fonts_drawing(xml_dir: str, fonts: dict[str, str]) -> int:
    """Add a:ea font element to a:rPr elements in DrawingML XML."""
    ea_font = fonts.get("eastAsia", "Malgun Gothic")
    count = 0
    for path in glob.glob(os.path.join(xml_dir, "**", "*.xml"), recursive=True):
        with open(path, "r", encoding="utf-8") as fh:
            original = fh.read()
        try:
            dom = defusedxml.minidom.parseString(original.encode("utf-8"))
        except Exception:
            continue

        modified = False
        for rpr in dom.getElementsByTagName("a:rPr"):
            # Check if a:ea already exists
            has_ea = any(
                getattr(c, "tagName", None) == "a:ea"
                for c in rpr.childNodes
            )
            if not has_ea:
                ea_elem = dom.createElement("a:ea")
                ea_elem.setAttribute("typeface", ea_font)
                rpr.appendChild(ea_elem)
                modified = True

        if modified:
            with open(path, "wb") as fh:
                fh.write(dom.toxml(encoding="UTF-8"))
            count += 1
    return count


# ---------------------------------------------------------------------------
# WCAG 2.1 contrast ratio
# ---------------------------------------------------------------------------


def _relative_luminance(hex_color: str) -> float:
    """Return WCAG 2.1 relative luminance for a 6-digit hex color."""
    hex_color = hex_color.lstrip("#")
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    channels = []
    for c in (r, g, b):
        s = c / 255.0
        channels.append(s / 12.92 if s <= 0.04045 else ((s + 0.055) / 1.055) ** 2.4)
    return 0.2126 * channels[0] + 0.7152 * channels[1] + 0.0722 * channels[2]


def check_contrast(fg_hex: str, bg_hex: str) -> float:
    """Return the WCAG 2.1 contrast ratio between two colors.

    A ratio >= 4.5 passes AA for normal text.
    A ratio >= 7.0 passes AAA.

    >>> round(check_contrast("FFFFFF", "000000"), 1)
    21.0
    """
    l1 = _relative_luminance(fg_hex)
    l2 = _relative_luminance(bg_hex)
    lighter, darker = max(l1, l2), min(l1, l2)
    return (lighter + 0.05) / (darker + 0.05)


# ---------------------------------------------------------------------------
# CJK text-box width estimation (for PptxGenJS / docx-js)
# ---------------------------------------------------------------------------


def _is_wide_codepoint(code: int) -> bool:
    """Return True if codepoint is CJK wide (hangul, CJK unified, fullwidth)."""
    return (
        (0x1100 <= code <= 0x11FF)
        or (0x3000 <= code <= 0x9FFF)
        or (0xAC00 <= code <= 0xD7AF)
        or (0xF900 <= code <= 0xFAFF)
        or (0xFE30 <= code <= 0xFE4F)
        or (0xFF00 <= code <= 0xFFEF)
    )


def estimate_text_width_inches(text: str, font_size_pt: float) -> float:
    """Estimate rendered width in inches for text with CJK characters.

    Uses an empirical metric of ~0.0104 inch per character unit per point,
    with a 0.9 correction factor for full-width characters.
    """
    char_units = 0
    for ch in text:
        code = ord(ch)
        char_units += 2 if _is_wide_codepoint(code) else 1
    avg_char_width = font_size_pt * 0.0104
    return char_units * avg_char_width * 0.9


# ---------------------------------------------------------------------------
# EMU helpers (OOXML image sizing)
# ---------------------------------------------------------------------------

EMU_PER_INCH = 914_400
EMU_PER_CM = 360_000


def pixel_to_emu(px: int, dpi: int = 96) -> int:
    """Convert pixels to English Metric Units."""
    return round(px * EMU_PER_INCH / dpi)


def emu_to_pixel(emu: int, dpi: int = 96) -> int:
    """Convert English Metric Units to pixels."""
    return round(emu * dpi / EMU_PER_INCH)
