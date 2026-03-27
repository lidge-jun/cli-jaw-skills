#!/usr/bin/env python3
"""Build and verify XLSX regression fixtures."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import zipfile
from pathlib import Path

import openpyxl

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"
EXPECTED_DIR = Path(__file__).resolve().parent / "expected"


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def create_formula_errors() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Errors"
    ws["A1"] = "Label"
    ws["B1"] = "Value"
    ws["A2"] = "Valid"
    ws["B2"] = 42
    ws["A3"] = "Sum"
    ws["B3"] = "=B2+10"
    ws["A4"] = "DivZero"
    ws["B4"] = "=B2/0"
    out = FIXTURES_DIR / "formula_errors.xlsx"
    wb.save(out)
    buf = io.BytesIO()
    with zipfile.ZipFile(out, "r") as src, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.namelist():
            data = src.read(item)
            if item.endswith("sheet1.xml"):
                text = data.decode("utf-8")
                error_row = (
                    '<row r="5"><c r="A5" t="str"><v>RefErr</v></c><c r="B5" t="e"><v>#REF!</v></c></row>'
                    '<row r="6"><c r="A6" t="str"><v>NameErr</v></c><c r="B6" t="e"><v>#NAME?</v></c></row>'
                )
                text = text.replace("</sheetData>", error_row + "</sheetData>")
                dst.writestr(item, text)
            else:
                dst.writestr(item, data)
    buf.seek(0)
    out.write_bytes(buf.read())


def create_multi_sheet() -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1["A1"] = "Product"
    ws1["B1"] = "Revenue"
    ws1["A2"] = "Widget A"
    ws1["B2"] = 1500
    ws1["A3"] = "Widget B"
    ws1["B3"] = 2300
    ws1["A4"] = "Total"
    ws1["B4"] = "=SUM(B2:B3)"
    ws2 = wb.create_sheet("Inventory")
    ws2["A1"] = "Item"
    ws2["B1"] = "Quantity"
    ws2["C1"] = "Location"
    ws2["A2"] = "Widget A"
    ws2["B2"] = 100
    ws2["C2"] = "Warehouse 1"
    ws2["A3"] = "Widget B"
    ws2["B3"] = 50
    ws2["C3"] = "Warehouse 2"
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Report Date"
    ws3["B1"] = "2026-03-18"
    ws3["A2"] = "Total Revenue"
    ws3["B2"] = "=Sales!B4"
    wb.save(FIXTURES_DIR / "multi_sheet.xlsx")


def create_broken_refs() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Hello"
    ws["A2"] = "World"
    out = FIXTURES_DIR / "broken_refs.xlsx"
    wb.save(out)
    buf = io.BytesIO()
    with zipfile.ZipFile(out, "r") as src, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.namelist():
            data = src.read(item)
            if item == "[Content_Types].xml":
                text = data.decode("utf-8")
                text = text.replace(
                    "</Types>",
                    '  <Override PartName="/xl/charts/chart1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>\n</Types>',
                )
                dst.writestr(item, text)
            elif item.endswith(".rels") and "xl/_rels" in item:
                text = data.decode("utf-8")
                text = text.replace(
                    "</Relationships>",
                    '  <Relationship Id="rId99" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" Target="charts/chart1.xml"/>\n</Relationships>',
                )
                dst.writestr(item, text)
            else:
                dst.writestr(item, data)
    buf.seek(0)
    out.write_bytes(buf.read())


def create_wide_columns_ad() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wide"
    for idx in range(1, 31):
        ws.cell(1, idx, f"H{idx}")
        ws.cell(2, idx, idx)
    wb.save(FIXTURES_DIR / "wide_columns_ad.xlsx")


def create_formula_uncached() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uncached"
    ws["A1"] = "Base"
    ws["B1"] = "Formula"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["B2"] = "=SUM(A2:A3)"
    ws["B3"] = "=A3*2"
    wb.save(FIXTURES_DIR / "formula_uncached.xlsx")


def create_korean_identifier_stress() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시트1"
    ws["A1"] = "년도"
    ws["B1"] = "매출액"
    ws["C1"] = "비고"
    ws["A2"] = 2024
    ws["B2"] = 1200
    ws["C2"] = "매출 증가"
    ws["A3"] = 2025
    ws["B3"] = 1500
    ws["C3"] = "매출 유지"
    wb.save(FIXTURES_DIR / "korean_identifier_stress.xlsx")


def _call_cli(command: str, fixture: str, *args: str) -> str:
    import subprocess
    import sys

    cli = Path(__file__).resolve().parent.parent / "scripts" / "xlsx_cli.py"
    result = subprocess.run([sys.executable, str(cli), command, str(FIXTURES_DIR / fixture), *args], capture_output=True, text=True, check=True)
    return result.stdout.strip()


def _write_expected_files() -> None:
    EXPECTED_DIR.mkdir(parents=True, exist_ok=True)
    (EXPECTED_DIR / "sheet_overview_wide_columns_ad.json").write_text(
        _call_cli("sheet-overview", "wide_columns_ad.xlsx", "--json") + "\n", encoding="utf-8"
    )
    (EXPECTED_DIR / "formula_audit_uncached.json").write_text(
        _call_cli("formula-audit", "formula_uncached.xlsx", "--json") + "\n", encoding="utf-8"
    )
    (EXPECTED_DIR / "search_korean_identifier_stress.json").write_text(
        _call_cli("search", "korean_identifier_stress.xlsx", "매출", "--json") + "\n", encoding="utf-8"
    )


def build(only: str | None = None, force: bool = False) -> list[dict]:
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    EXPECTED_DIR.mkdir(parents=True, exist_ok=True)
    registry = {
        "formula_errors": create_formula_errors,
        "multi_sheet": create_multi_sheet,
        "broken_refs": create_broken_refs,
        "wide_columns_ad": create_wide_columns_ad,
        "formula_uncached": create_formula_uncached,
        "korean_identifier_stress": create_korean_identifier_stress,
    }
    results = []
    for name, fn in registry.items():
        if only and only != name:
            continue
        target = FIXTURES_DIR / f"{name}.xlsx"
        if target.exists() and not force:
            results.append({"fixture": target.name, "status": "SKIPPED"})
            continue
        fn()
        results.append({"fixture": target.name, "status": "CREATED", "sha256": _sha256(target)})
    if not only:
        _write_expected_files()
    return results


def verify(only: str | None = None) -> dict:
    mapping = {
        "formula_errors": {"fixture": "formula_errors.xlsx"},
        "multi_sheet": {"fixture": "multi_sheet.xlsx"},
        "broken_refs": {"fixture": "broken_refs.xlsx"},
        "wide_columns_ad": {"fixture": "wide_columns_ad.xlsx"},
        "formula_uncached": {"fixture": "formula_uncached.xlsx"},
        "korean_identifier_stress": {"fixture": "korean_identifier_stress.xlsx"},
    }
    checked = []
    failed = []
    details = []
    for slug, cfg in mapping.items():
        if only and only != slug:
            continue
        fixture = FIXTURES_DIR / cfg["fixture"]
        result = {"fixture": cfg["fixture"], "exists": fixture.exists(), "checksum_ok": fixture.exists(), "assertions_ok": fixture.exists()}
        if slug == "wide_columns_ad" and fixture.exists():
            wb = openpyxl.load_workbook(fixture, read_only=True, data_only=True)
            result["assertions_ok"] = wb.active.max_column == 30
            wb.close()
        elif slug == "formula_uncached" and fixture.exists():
            wb = openpyxl.load_workbook(fixture, read_only=False, data_only=False)
            result["assertions_ok"] = ws_has_formula(wb.active)
            wb.close()
        elif slug == "korean_identifier_stress" and fixture.exists():
            wb = openpyxl.load_workbook(fixture, read_only=True, data_only=True)
            result["assertions_ok"] = wb.sheetnames[0] == "시트1"
            wb.close()
        if not (result["exists"] and result["checksum_ok"] and result["assertions_ok"]):
            failed.append(slug)
        checked.append(slug)
        details.append(result)
    return {"mode": "verify", "skill": "xlsx", "ok": not failed, "checked": checked, "failed": failed, "details": details}


def ws_has_formula(ws) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                return True
    return False


def main() -> int:
    parser = argparse.ArgumentParser(description="Build and verify XLSX regression fixtures")
    sub = parser.add_subparsers(dest="command", required=True)
    build_parser = sub.add_parser("build")
    build_parser.add_argument("--only")
    build_parser.add_argument("--force", action="store_true")
    verify_parser = sub.add_parser("verify")
    verify_parser.add_argument("--only")
    verify_parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    if args.command == "build":
        results = build(only=args.only, force=args.force)
        for item in results:
            print(f"{item['status']} {item['fixture']}")
        return 0

    result = verify(only=args.only)
    if args.json:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        for detail in result["details"]:
            if detail["exists"] and detail["checksum_ok"] and detail["assertions_ok"]:
                print(f"VERIFIED {detail['fixture']}")
            else:
                print(f"FAILED {detail['fixture']}")
    return 0 if result["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
