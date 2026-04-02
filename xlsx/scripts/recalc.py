"""Recalculate formulas in an XLSX file via LibreOffice.

Opens the spreadsheet in LibreOffice headless mode, forces formula
recalculation, saves, then optionally scans for formula errors.

Usage:
    python recalc.py <input.xlsx> <output.xlsx>
    python recalc.py <input.xlsx> <output.xlsx> --check-errors
"""

import argparse
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
SKILLS_REF_DIR = SCRIPT_DIR.parent.parent

import sys
sys.path.insert(0, str(SKILLS_REF_DIR))
sys.path.insert(0, str(SCRIPT_DIR))

try:
    from ooxml_core.soffice import run_soffice
except ImportError:
    from ooxml.soffice import run_soffice

# Common Excel formula error values
FORMULA_ERRORS = {
    "#REF!",
    "#NAME?",
    "#VALUE!",
    "#DIV/0!",
    "#NULL!",
    "#N/A",
    "#NUM!",
    "#CALC!",
    "#SPILL!",
    "#GETTING_DATA",
}


def recalc(input_path: str, output_path: str, check_errors: bool = False) -> str:
    """Recalculate all formulas in an XLSX file."""
    src = Path(input_path).resolve()
    dst = Path(output_path).resolve()

    if not src.exists():
        return f"Error: {input_path} does not exist"

    try:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tmp_file = tmp_path / src.name
            shutil.copy2(src, tmp_file)

            # Run LibreOffice macro to recalculate
            result = run_soffice(
                [
                    "--headless",
                    "--calc",
                    "--convert-to", "xlsx",
                    "--outdir", str(tmp_path),
                    str(tmp_file),
                ],
                capture_output=True,
                text=True,
                timeout=120,
            )

            # Find the output file
            output_files = list(tmp_path.glob("*.xlsx"))
            if not output_files:
                return f"Error: recalculation failed. stderr: {result.stderr}"

            # Copy to destination
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(output_files[0], dst)

        msg = f"Recalculated formulas: {input_path} → {output_path}"

        if check_errors:
            errors = scan_formula_errors(str(dst))
            if errors:
                msg += f"\n\nFormula errors found ({len(errors)}):"
                for err in errors[:20]:  # limit output
                    msg += f"\n  - {err}"
                if len(errors) > 20:
                    msg += f"\n  ... and {len(errors) - 20} more"

        return msg

    except Exception as e:
        return f"Error: {e}"


def scan_formula_errors(xlsx_path: str) -> list[str]:
    """Scan an XLSX file for formula error values using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl not installed — cannot scan for errors"]

    errors = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        if cell.value in FORMULA_ERRORS:
                            coord = f"{sheet_name}!{cell.coordinate}"
                            errors.append(f"{coord}: {cell.value}")

        wb.close()
    except Exception as e:
        errors.append(f"Scan error: {e}")

    return errors


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Recalculate formulas in an XLSX file"
    )
    parser.add_argument("input_file", help="Input XLSX with formulas")
    parser.add_argument("output_file", help="Output XLSX with recalculated values")
    parser.add_argument(
        "--check-errors",
        action="store_true",
        help="Scan for formula error values after recalculation",
    )
    args = parser.parse_args()

    msg = recalc(args.input_file, args.output_file, check_errors=args.check_errors)
    print(msg)
    sys.exit(1 if msg.startswith("Error") else 0)
