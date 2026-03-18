#!/usr/bin/env python3
"""Unified CLI for XLSX document operations.

Subcommands:
    open            Unpack XLSX to directory with pretty-printed XML
    save            Pack directory to XLSX
    validate        Structural validation (JSON output)
    repair          Auto-repair common issues (dry-run by default)
    recalc          Recalculate formulas via LibreOffice
    text            Extract cell text from XLSX
    sheet-overview  List sheets with row/column counts
    formula-audit   Scan for formula errors across all sheets
    search          Search text/values across sheets

Usage:
    python xlsx_cli.py open input.xlsx work/
    python xlsx_cli.py save work/ output.xlsx
    python xlsx_cli.py validate input.xlsx
    python xlsx_cli.py validate input.xlsx --json
    python xlsx_cli.py repair input.xlsx
    python xlsx_cli.py repair input.xlsx --apply -o out.xlsx
    python xlsx_cli.py recalc input.xlsx output.xlsx
    python xlsx_cli.py recalc input.xlsx output.xlsx --check-errors
    python xlsx_cli.py text input.xlsx
    python xlsx_cli.py sheet-overview input.xlsx
    python xlsx_cli.py formula-audit input.xlsx
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _unpack_to_tmpdir(xlsx_path: str) -> str:
    """Unpack XLSX to a temp directory, return path."""
    import zipfile
    tmpdir = tempfile.mkdtemp(prefix="xlsx_cli_")
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        zf.extractall(tmpdir)
    return tmpdir


def _load_workbook_readonly(xlsx_path: str):
    """Load workbook in read-only mode for inspection."""
    import openpyxl
    return openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)


# ---------------------------------------------------------------------------
# Subcommands
# ---------------------------------------------------------------------------

def cmd_open(args: argparse.Namespace) -> int:
    """Unpack XLSX with pretty-printed XML."""
    sys.path.insert(0, str(SCRIPT_DIR))
    from ooxml.unpack import unpack
    msg = unpack(args.input, args.output_dir)
    print(msg)
    return 1 if msg.startswith("Error") else 0


def cmd_save(args: argparse.Namespace) -> int:
    """Pack directory to XLSX."""
    sys.path.insert(0, str(SCRIPT_DIR))
    from ooxml.pack import pack
    msg = pack(args.input_dir, args.output)
    print(msg)
    return 1 if msg.startswith("Error") else 0


def cmd_validate(args: argparse.Namespace) -> int:
    """Structural validation."""
    sys.path.insert(0, str(SCRIPT_DIR))
    from ooxml.validate import validate
    result = validate(args.input, verbose=not args.json)
    if args.json:
        print(json.dumps(result, indent=2))
    else:
        if not result["passed"]:
            for e in result["errors"]:
                print(f"  ERROR: {e}")
            for w in result["warnings"]:
                print(f"  WARN:  {w}")
    return 0 if result["passed"] else 1


def cmd_repair(args: argparse.Namespace) -> int:
    """Auto-repair common OOXML issues (dry-run by default)."""
    apply = args.apply or args.output is not None
    dry_run = not apply

    work = _unpack_to_tmpdir(args.input)
    try:
        sys.path.insert(0, str(SCRIPT_DIR))
        from ooxml.repair import repair
        result = repair(work, dry_run=dry_run)

        if args.json:
            print(json.dumps(result, indent=2))
        else:
            if result["repaired"] > 0:
                label = "REPAIRED" if apply else "WOULD REPAIR"
                print(f"{label} ({result['repaired']} issue(s)):")
                for d in result["details"]:
                    print(f"  - {d}")
            else:
                print("No issues found.")

        if apply and result["repaired"] > 0:
            from ooxml.pack import pack
            output = args.output or args.input
            if not args.output:
                backup = Path(args.input).with_suffix(".xlsx.bak")
                if not backup.exists():
                    shutil.copy2(args.input, backup)
                    print(f"Backup: {backup}")
            pack(work, output)
            print(f"Output: {output}")

            from ooxml.validate import validate
            vresult = validate(output)
            if vresult["passed"]:
                print("Post-repair validation: PASS")
            else:
                print("WARNING: post-repair validation issues:", file=sys.stderr)
                for e in vresult["errors"]:
                    print(f"  - {e}", file=sys.stderr)

        return 0
    finally:
        shutil.rmtree(work, ignore_errors=True)


def cmd_recalc(args: argparse.Namespace) -> int:
    """Recalculate formulas via LibreOffice."""
    sys.path.insert(0, str(SCRIPT_DIR))
    from recalc import recalc
    msg = recalc(args.input, args.output, check_errors=args.check_errors)
    print(msg)
    return 1 if msg.startswith("Error") else 0


def cmd_text(args: argparse.Namespace) -> int:
    """Extract cell text from XLSX."""
    try:
        wb = _load_workbook_readonly(args.input)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"--- {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                values = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in values):
                    print("\t".join(values))
    finally:
        wb.close()
    return 0


def cmd_sheet_overview(args: argparse.Namespace) -> int:
    """List sheets with dimensions and summary."""
    try:
        wb = _load_workbook_readonly(args.input)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    try:
        overview = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            info = {
                "name": sheet_name,
                "dimensions": getattr(ws, "dimensions", None) or "empty",
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }
            overview.append(info)

        if args.json:
            print(json.dumps(overview, indent=2, ensure_ascii=False))
        else:
            print(f"Workbook: {args.input} ({len(overview)} sheet(s))\n")
            for s in overview:
                print(f"  [{s['name']}] {s['dimensions']}  ({s['max_row']} rows x {s['max_column']} cols)")
    finally:
        wb.close()
    return 0


def cmd_formula_audit(args: argparse.Namespace) -> int:
    """Scan for formula errors across all sheets."""
    sys.path.insert(0, str(SCRIPT_DIR))
    from recalc import scan_formula_errors, FORMULA_ERRORS

    errors = scan_formula_errors(args.input)

    if args.json:
        result = {"passed": len(errors) == 0, "errors": errors}
        print(json.dumps(result, indent=2))
    else:
        if errors:
            print(f"Formula errors found ({len(errors)}):")
            for err in errors:
                print(f"  - {err}")
        else:
            print("No formula errors found.")

    return 1 if errors else 0


def cmd_search(args: argparse.Namespace) -> int:
    """Search text/values across sheets with regex."""
    import re

    pattern = re.compile(args.pattern)
    try:
        wb = _load_workbook_readonly(args.input)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    results: list[dict] = []
    try:
        sheets = [args.sheet] if args.sheet else wb.sheetnames
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                print(f"Error: Sheet '{sheet_name}' not found.", file=sys.stderr)
                wb.close()
                return 1
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    if pattern.search(text):
                        results.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "text": text[:120],
                        })
    finally:
        wb.close()

    if not results:
        print("No matches found.", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(results, ensure_ascii=False, indent=2))
    else:
        for r in results:
            print(f"{r['sheet']}:{r['cell']}: {r['text']}")
    print(f"\n{len(results)} match(es) found.", file=sys.stderr)
    return 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Unified CLI for XLSX document operations",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # open
    p = sub.add_parser("open", help="Unpack XLSX to directory with pretty-printed XML")
    p.add_argument("input", help="Input .xlsx file")
    p.add_argument("output_dir", help="Output directory")

    # save
    p = sub.add_parser("save", help="Pack directory to XLSX")
    p.add_argument("input_dir", help="Input directory")
    p.add_argument("output", help="Output .xlsx file")

    # validate
    p = sub.add_parser("validate", help="Structural validation")
    p.add_argument("input", help="Input .xlsx file or unpacked directory")
    p.add_argument("--json", "-j", action="store_true", help="JSON output")

    # repair
    p = sub.add_parser("repair", help="Auto-repair (dry-run by default)")
    p.add_argument("input", help="Input .xlsx file")
    p.add_argument("--apply", action="store_true", help="Apply repairs (default: dry-run)")
    p.add_argument("-o", "--output", help="Output file (implies --apply)")
    p.add_argument("--json", "-j", action="store_true", help="JSON output")

    # recalc
    p = sub.add_parser("recalc", help="Recalculate formulas via LibreOffice")
    p.add_argument("input", help="Input .xlsx file")
    p.add_argument("output", help="Output .xlsx file")
    p.add_argument("--check-errors", action="store_true", help="Scan for errors after recalc")

    # text
    p = sub.add_parser("text", help="Extract cell text from XLSX")
    p.add_argument("input", help="Input .xlsx file")

    # sheet-overview
    p = sub.add_parser("sheet-overview", help="List sheets with dimensions")
    p.add_argument("input", help="Input .xlsx file")
    p.add_argument("--json", "-j", action="store_true", help="JSON output")

    # formula-audit
    p = sub.add_parser("formula-audit", help="Scan for formula errors")
    p.add_argument("input", help="Input .xlsx file")
    p.add_argument("--json", "-j", action="store_true", help="JSON output")

    # search
    p = sub.add_parser("search", help="Search text/values across sheets")
    p.add_argument("input", help="Input .xlsx file")
    p.add_argument("pattern", help="Regex pattern to search")
    p.add_argument("--sheet", help="Limit search to specific sheet name")
    p.add_argument("--json", "-j", action="store_true", help="JSON output")

    args = parser.parse_args()

    commands = {
        "open": cmd_open,
        "save": cmd_save,
        "validate": cmd_validate,
        "repair": cmd_repair,
        "recalc": cmd_recalc,
        "text": cmd_text,
        "sheet-overview": cmd_sheet_overview,
        "formula-audit": cmd_formula_audit,
        "search": cmd_search,
    }

    return commands[args.command](args)


if __name__ == "__main__":
    raise SystemExit(main())
